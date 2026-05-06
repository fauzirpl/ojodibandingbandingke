from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import uuid
import json
import numpy as np
from pipeline import FuzzyMatcherPipeline
from ai_validator import AIValidator
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="Fuzzy Matcher API")

# Global cache for intermediate results (In production, use Redis)
RESULTS_CACHE = {}
@app.get("/config")
async def get_config():
    return {
        "ai_model": os.getenv("AI_MODEL", "google/gemini-2.0-flash-001")
    }

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "temp_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Store dataframes in memory for selection (simple session-like behavior)
# In production, use Redis or a database
sessions = {}

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    session_id = str(uuid.uuid4())
    file_ext = os.path.splitext(file.filename)[1].lower()
    file_path = os.path.join(UPLOAD_DIR, f"{session_id}{file_ext}")
    
    with open(file_path, "wb") as buffer:
        buffer.write(await file.read())
    
    try:
        if file_ext == '.csv':
            df = pd.read_csv(file_path, nrows=5) # Just preview
            all_cols = pd.read_csv(file_path, nrows=0).columns.tolist()
        else:
            df = pd.read_excel(file_path, nrows=5)
            all_cols = pd.read_excel(file_path, nrows=0).columns.tolist()
            
        return {
            "session_id": session_id,
            "columns": all_cols,
            "preview": df.replace({np.nan: None}).to_dict('records'),
            "filename": file.filename
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/match")
async def start_match(
    session_left: str = Form(...),
    session_right: str = Form(...),
    cols_left: str = Form(...),
    cols_right: str = Form(...),
    threshold: float = Form(30.0),
    export_cols: str = Form(None),
    ai_config: str = Form(None),
    use_ai: bool = Form(False)
):
    try:
        # Load AI config from .env
        ai_key = os.getenv("API_KEY")
        ai_model = os.getenv("AI_MODEL", "google/gemini-2.0-flash-001")
        
        # Parse columns
        c_left = json.loads(cols_left)
        c_right = json.loads(cols_right)
        e_cols = json.loads(export_cols) if export_cols else None
        a_config = json.loads(ai_config) if ai_config else None
        
        # Find files
        file_left = None
        file_right = None
        for f in os.listdir(UPLOAD_DIR):
            if f.startswith(session_left):
                file_left = os.path.join(UPLOAD_DIR, f)
            if f.startswith(session_right):
                file_right = os.path.join(UPLOAD_DIR, f)
        
        if not file_left or not file_right:
            raise HTTPException(status_code=404, detail="Files not found")
            
        def load_df(path):
            ext = os.path.splitext(path)[1].lower()
            return pd.read_csv(path) if ext == '.csv' else pd.read_excel(path)
            
        df_a = load_df(file_left)
        df_b = load_df(file_right)
        
        # Initialize AI Validator
        ai_validator = None
        if ai_key:
            ai_validator = AIValidator(api_key=ai_key, model=ai_model)

        matcher = FuzzyMatcherPipeline(threshold=threshold, ai_validator=ai_validator)
        result_df = matcher.run(df_a, df_b, c_left, c_right, ai_config=a_config, use_ai=use_ai)
        
        # Store in cache for Step 2
        result_id = str(uuid.uuid4())
        RESULTS_CACHE[result_id] = {
            "df": result_df,
            "export_cols": e_cols,
            "ai_config": a_config,
            "matcher": matcher
        }
        
        return {
            "result_id": result_id,
            "total_rows": len(result_df),
            "preview": result_df.head(50).replace({np.nan: None}).to_dict('records'),
            "is_ai_complete": use_ai
        }
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/verify-ai/{result_id}")
async def verify_ai(result_id: str):
    # Keep this for backward compatibility or bulk processing if needed
    if result_id not in RESULTS_CACHE:
        raise HTTPException(status_code=404, detail="Result session expired or not found")
    
    cached = RESULTS_CACHE[result_id]
    df = cached["df"]
    matcher = cached["matcher"]
    ai_config = cached["ai_config"]
    
    # Filter: Only score >= 50
    # The logic is now inside run_ai_reranking or controlled by caller
    updated_df = matcher.run_ai_reranking(df, ai_config)
    cached["df"] = updated_df
    
    return {
        "result_id": result_id,
        "total_rows": len(updated_df),
        "preview": updated_df.head(50).replace({np.nan: None}).to_dict('records'),
        "is_ai_complete": True
    }

@app.post("/verify-chunk")
async def verify_chunk(
    pairs: list = Form(...),
    ai_config: str = Form(None)
):
    """Secure proxy for AI verification of a small chunk of pairs."""
    try:
        ai_key = os.getenv("API_KEY")
        ai_model = os.getenv("AI_MODEL", "google/gemini-2.0-flash-001")
        
        if not ai_key:
            raise HTTPException(status_code=500, detail="API Key not configured")
            
        validator = AIValidator(api_key=ai_key, model=ai_model)
        chunk_data = json.loads(pairs)
        config_data = json.loads(ai_config) if ai_config else None
        
        results = validator._process_chunk(chunk_data, config_data)
        return results
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{result_id}")
async def download_result(result_id: str):
    if result_id not in RESULTS_CACHE:
        # Check disk fallback
        result_path = os.path.join(UPLOAD_DIR, f"result_{result_id}.xlsx")
        if os.path.exists(result_path):
            return FileResponse(result_path, filename=f"matching_result_{result_id[:8]}.xlsx")
        raise HTTPException(status_code=404, detail="Result not found")
    
    cached = RESULTS_CACHE[result_id]
    result_df = cached["df"]
    e_cols = cached["export_cols"]
    
    # Filter columns and ensure internal columns are hidden
    available_cols = [c for c in result_df.columns if not c.startswith('_')]
    valid_e_cols = [c for c in e_cols if c in available_cols] if e_cols else available_cols
    export_df = result_df[valid_e_cols]
    
    # Generate Excel
    result_path = os.path.join(UPLOAD_DIR, f"result_{result_id}.xlsx")
    
    pretty_df = export_df.copy()
    new_cols = {}
    for col in pretty_df.columns:
        if col.endswith('_left'):
            new_cols[col] = f"[A] {col.replace('_left', '')}"
        elif col.endswith('_right'):
            new_cols[col] = f"[B] {col.replace('_right', '')}"
        else:
            name = col.replace('_', ' ').title()
            name = name.replace('Ai', 'AI')
            new_cols[col] = name
    pretty_df.rename(columns=new_cols, inplace=True)
    
    pretty_df.to_excel(result_path, index=False, engine='openpyxl')
    
    # Styling
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = load_workbook(result_path)
    ws = wb.active
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    wb.save(result_path)
    return FileResponse(result_path, filename=f"matching_result_{result_id[:8]}.xlsx")

@app.get("/all-candidates/{result_id}")
async def get_all_candidates(result_id: str):
    """Returns ALL candidates with score >= 50 for full AI processing."""
    if result_id not in RESULTS_CACHE:
        raise HTTPException(status_code=404, detail="Session expired")
    
    df = RESULTS_CACHE[result_id]["df"]
    # Filter only potential candidates for AI
    ai_candidates = df[df['score'] >= 50].copy()
    
    return ai_candidates.replace({np.nan: None}).to_dict('records')

@app.post("/export-custom")
async def export_custom(
    result_id: str = Form(None),
    data: str = Form(...),
    columns: str = Form(...)
):
    try:
        if not result_id or result_id not in RESULTS_CACHE:
            raise HTTPException(status_code=404, detail="Sesi matching tidak ditemukan atau sudah kadaluarsa")
            
        # 1. Ambil data asli dari cache
        cached = RESULTS_CACHE[result_id]
        full_df = cached["df"].copy()
        e_cols = cached["export_cols"]
        
        # 2. Ambil hasil AI dari frontend dan masukkan ke full_df
        ai_updates = json.loads(data)
        ai_df = pd.DataFrame(ai_updates)
        
        if not ai_df.empty and '_row_id' in ai_df.columns:
            # Update status AI di data asli berdasarkan _row_id
            for _, row in ai_df.iterrows():
                idx = full_df[full_df['_row_id'] == row['_row_id']].index
                if not idx.empty:
                    for col in ['ai_status', 'ai_score', 'ai_reason']:
                        if col in row:
                            full_df.loc[idx, col] = row[col]

        # 3. Logika Deduplikasi (Ambil 1 terbaik per _row_id)
        # Prioritas: COCOK (3) > PERLU_VERIFIKASI (2) > TIDAK_COCOK (1) > Lainnya (0)
        priority_map = {'COCOK': 3, 'PERLU_VERIFIKASI': 2, 'TIDAK_COCOK': 1}
        full_df['ai_priority'] = full_df['ai_status'].map(lambda x: priority_map.get(x, 0))
        
        # Sortir berdasarkan _row_id, lalu prioritas AI, lalu skor semantik
        full_df = full_df.sort_values(
            by=['_row_id', 'ai_priority', 'ai_score', 'score'], 
            ascending=[True, False, False, False]
        )
        
        # Ambil baris pertama untuk setiap _row_id (yang terbaik)
        full_df = full_df.drop_duplicates(subset=['_row_id'], keep='first')

        # 4. Logika Filter & Urutan (PERSIS download_result)
        available_cols = [c for c in full_df.columns if not str(c).startswith('_')]
        
        # Tambahkan kolom AI ke daftar kolom yang tersedia jika belum ada
        ai_cols = ['score', 'ai_score', 'ai_status', 'ai_reason']
        
        if e_cols:
            # Jika user pilih kolom tertentu, pastikan kolom AI ditambahkan di akhir
            valid_cols = [c for c in e_cols if c in available_cols]
            # Tambahkan score dan ai_cols jika belum masuk
            for c in ai_cols:
                if c in available_cols and c not in valid_cols:
                    valid_cols.append(c)
        else:
            valid_cols = available_cols

        export_df = full_df[valid_cols].copy()
        
        # 4. Penamaan Kolom (PERSIS download_result)
        new_cols = {}
        for col in export_df.columns:
            if col.endswith('_left'):
                new_cols[col] = f"[A] {col.replace('_left', '')}"
            elif col.endswith('_right'):
                new_cols[col] = f"[B] {col.replace('_right', '')}"
            else:
                name = col.replace('_', ' ').title().replace('Ai', 'AI')
                new_cols[col] = name
        
        export_df.rename(columns=new_cols, inplace=True)
        export_df = export_df.replace({np.nan: None})
        
        # 5. Generate Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Hasil_Matching')
            ws = writer.sheets['Hasil_Matching']
            
            # Styling (Sama dengan awal)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=hasil_matching_ai_final.xlsx"}
        )
    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
